import streamlit as st
import pandas as pd
import tempfile
import os
import re
from typing import List, Dict, Tuple
import pdfplumber
from openpyxl import Workbook

st.set_page_config(page_title="PDF Extractor", page_icon="📊", layout="wide")
st.title("📊 PDF Extractor")

def extract_pdf(pdf_path: str):
    """Extract data from PDF"""
    rows = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                lines = page.extract_text().split('\n')
                
                for i, line in enumerate(lines):
                    # Look for invoice numbers
                    if re.search(r'\b9[0-9]{7,8}\b', line):
                        try:
                            invoice_match = re.search(r'\b(9[0-9]{7,8})\b', line)
                            invoice_num = invoice_match.group(1)
                            
                            # Extract numbers from line
                            numbers = re.findall(r'-?\d+\.?\d*', line)
                            if len(numbers) >= 2:
                                commission = float(numbers[-1])
                                unit_cost = float(numbers[-3]) if len(numbers) >= 3 else 0
                                
                                # Get customer name
                                cust_name = line[:invoice_match.start()].strip()
                                
                                # Get next lines
                                city, state, po = "", "", ""
                                cust_acc = None
                                
                                if i+1 < len(lines):
                                    second_line = lines[i+1].strip()
                                    state_match = re.search(r'\b([A-Z]{2})\b', second_line)
                                    if state_match:
                                        state = state_match.group(1)
                                
                                if i+2 < len(lines):
                                    third_line = lines[i+2].strip()
                                    acc_match = re.search(r'\b(\d{5,6})\b', third_line)
                                    if acc_match:
                                        cust_acc = int(acc_match.group(1))
                                
                                row = ('Prime', None, cust_name, city, state, cust_acc, invoice_num, po, unit_cost, 1, commission)
                                rows.append(row)
                        except:
                            continue
        
        return rows
    except Exception as e:
        st.error(f"Error: {str(e)}")
        return []

# Upload
uploaded_files = st.file_uploader("Upload PDF", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    if st.button("Extract & Download"):
        st.info("Processing...")
        
        for file in uploaded_files:
            # Save temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(file.getbuffer())
                tmp_path = tmp.name
            
            # Extract
            rows = extract_pdf(tmp_path)
            os.unlink(tmp_path)
            
            if rows:
                st.success(f"✅ {file.name}: {len(rows)} rows extracted")
                
                # Create Excel
                wb = Workbook()
                ws = wb.active
                headers = ['Supplier_name', 'Distname', 'CustName', 'City', 'State', 'CustAccNbr', 'InvoiceNumber', 'PO_Number', 'UnitCost', 'Qty', 'Commissions']
                
                for col, header in enumerate(headers, 1):
                    ws.cell(1, col, header)
                
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row_idx, col_idx, val)
                
                excel_file = f"export_{file.name.split('.')[0]}.xlsx"
                wb.save(excel_file)
                
                # Download button
                with open(excel_file, 'rb') as f:
                    st.download_button(
                        label=f"⬇️ Download {file.name.split('.')[0]}.xlsx",
                        data=f.read(),
                        file_name=excel_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                os.unlink(excel_file)
            else:
                st.error(f"❌ {file.name}: No data extracted")
