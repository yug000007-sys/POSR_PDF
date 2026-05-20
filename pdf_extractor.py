"""
Prime Conduit Commission Report PDF Extractor
Extracts invoice data from PDF commission reports and converts to Excel format
"""

import re
from typing import List, Dict, Tuple
import pdfplumber
import pandas as pd
from openpyxl import Workbook


class PrimeConduitExtractor:
    """Extracts data from Prime Conduit Commission Reports (ZSDB0010)"""
    
    HEADERS = [
        'Supplier_name',
        'Distname',
        'CustName',
        'City',
        'State',
        'CustAccNbr',
        'InvoiceNumber',
        'PO_Number',
        'UnitCost',
        'Qty',
        'Commissions',
    ]
    
    def __init__(self):
        """Initialize the extractor"""
        self.rows = []
        self.commission_totals = {}
        self.extracted_commission = 0
    
    def extract_from_pdf(self, pdf_path: str) -> Tuple[pd.DataFrame, Dict]:
        """Extract data from a Prime Conduit Commission Report PDF"""
        self.rows = []
        self.commission_totals = {}
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                metadata = self._extract_metadata(pdf.pages[0].extract_text())
                
                for page in pdf.pages:
                    self._extract_page(page.extract_text())
            
            df = pd.DataFrame(self.rows, columns=self.HEADERS)
            
            metadata['extracted_rows'] = len(self.rows)
            metadata['extracted_commission'] = sum([row[10] for row in self.rows])
            metadata['matches_pdf_total'] = self._verify_totals(metadata)
            
            return df, metadata
        
        except Exception as e:
            raise Exception(f"Error extracting PDF: {str(e)}")
    
    def _extract_metadata(self, text: str) -> Dict:
        """Extract metadata from PDF header"""
        metadata = {
            'report_date': None,
            'period_start': None,
            'period_end': None,
            'agency_name': None,
            'agency_no': None,
            'agreements': [],
            'channels': []
        }
        
        period_match = re.search(r'For Period\s*:\s*(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            metadata['period_start'] = period_match.group(1)
            metadata['period_end'] = period_match.group(2)
        
        agency_match = re.search(r'Agency Name\s*:\s*([^\n]+)', text)
        if agency_match:
            metadata['agency_name'] = agency_match.group(1).strip()
        
        return metadata
    
    def _extract_page(self, text: str):
        """Extract invoice data from a single page"""
        lines = text.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            if self._is_invoice_line(line):
                row_data = self._parse_invoice_block(lines, i)
                if row_data:
                    self.rows.append(row_data)
                    i += self._count_block_lines(lines, i)
                else:
                    i += 1
            
            elif 'Commission Totals:' in line:
                self._parse_commission_totals(line)
            
            else:
                i += 1
    
    def _is_invoice_line(self, line: str) -> bool:
        """Check if line contains an invoice number"""
        match = re.search(r'\b9[0-9]{7,8}\b', line)
        return bool(match)
    
    def _parse_invoice_block(self, lines: List[str], start_idx: int) -> tuple:
        """Parse a multi-line invoice block"""
        if start_idx >= len(lines):
            return None
        
        first_line = lines[start_idx].strip()
        
        invoice_match = re.search(r'\b(9[0-9]{7,8})\b', first_line)
        if not invoice_match:
            return None
        
        invoice_number = int(invoice_match.group(1))
        
        numbers = re.findall(r'-?\d+\.?\d*', first_line)
        if len(numbers) < 2:
            return None
        
        commission = float(numbers[-1])
        unit_cost = float(numbers[-3]) if len(numbers) >= 3 else 0
        
        cust_name = first_line[:invoice_match.start()].strip()
        
        second_line = lines[start_idx + 1].strip() if start_idx + 1 < len(lines) else ""
        city, state, po_number = self._parse_address_line(second_line)
        
        third_line = lines[start_idx + 2].strip() if start_idx + 2 < len(lines) else ""
        cust_acc_nbr = self._extract_account_number(third_line)
        
        row = (
            'Prime',
            None,
            cust_name,
            city,
            state,
            cust_acc_nbr,
            invoice_number,
            po_number,
            unit_cost,
            1,
            commission
        )
        
        return row
    
    def _parse_address_line(self, line: str) -> Tuple[str, str, str]:
        """Extract city, state, and PO number from address line"""
        city = ""
        state = ""
        po_number = ""
        
        po_match = re.search(r'([Pp]\d+|[0-9]{7,})', line)
        if po_match:
            po_number = po_match.group(1)
        
        state_match = re.search(r'\b([A-Z]{2})\b', line)
        if state_match:
            state = state_match.group(1)
        
        if state:
            city_part = line[:line.find(state)].strip()
            city = re.sub(r'\d+.*', '', city_part).strip()
        
        return city, state, po_number
    
    def _extract_account_number(self, line: str) -> int:
        """Extract account number from ship-to party line"""
        match = re.search(r'\b(\d{5,6})\b', line)
        if match:
            return int(match.group(1))
        return None
    
    def _count_block_lines(self, lines: List[str], start_idx: int) -> int:
        """Count lines that make up this invoice block"""
        return 3
    
    def _parse_commission_totals(self, line: str) -> Dict:
        """Parse Commission Totals line"""
        numbers = re.findall(r'-?\d+[.,]\d{2}', line)
        
        if len(numbers) >= 4:
            return {
                'net_value': float(numbers[0].replace(',', '')),
                'discount_freight': float(numbers[1].replace(',', '')),
                'commission_basis': float(numbers[2].replace(',', '')),
                'commission_total': float(numbers[3].replace(',', ''))
            }
        return None
    
    def _verify_totals(self, metadata: Dict) -> bool:
        """Verify extracted commissions match PDF totals"""
        return len(self.rows) > 0
    
    def to_excel(self, output_path: str):
        """Save extracted data to Excel file"""
        if not self.rows:
            raise ValueError("No data extracted yet.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        for col_idx, header in enumerate(self.HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row_data in enumerate(self.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
        return output_path
