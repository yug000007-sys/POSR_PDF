import re
from io import BytesIO

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Prime Conduit PDF Extractor",
    page_icon="📊",
    layout="wide"
)

HEADERS = [
    "Supplier_name",
    "Distname",
    "CustName",
    "City",
    "State",
    "CustAccNbr",
    "InvoiceNumber",
    "PO_Number",
    "UnitCost",
    "Qty",
    "Commissions",
]

INVOICE_RE = re.compile(r"\b9\d{7,8}\b")
STATE_RE = re.compile(r"\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY)\b")
ACCOUNT_RE = re.compile(r"\b\d{5,6}\b")
MONEY_RE = re.compile(r"\$?\s*-?\d{1,3}(?:,\d{3})*(?:\.\d{2})")


def clean_text(value):
    if not value:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def money_to_number(value):
    if not value:
        return ""
    value = value.replace("$", "").replace(",", "").strip()
    try:
        return float(value)
    except ValueError:
        return value


def extract_rows_from_text(text):
    rows = []
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]

    for i, line in enumerate(lines):
        invoices = INVOICE_RE.findall(line)
        if not invoices:
            continue

        context_lines = lines[max(0, i - 3): min(len(lines), i + 4)]
        context = " ".join(context_lines)

        account_match = ACCOUNT_RE.search(context)
        state_match = STATE_RE.search(context)
        money_matches = MONEY_RE.findall(context)

        city = ""
        state = state_match.group(1) if state_match else ""

        if state:
            before_state = context[:context.find(state)].strip()
            city_parts = before_state.split()
            if city_parts:
                city = city_parts[-1]

        cust_name = ""
        for ctx_line in context_lines:
            if not INVOICE_RE.search(ctx_line):
                if any(word in ctx_line.upper() for word in ["ECHO", "BORDER", "GRAYBAR", "CED", "VIKING", "ELECTRIC", "SUPPLY"]):
                    cust_name = ctx_line
                    break

        if not cust_name and context_lines:
            cust_name = context_lines[0]

        commission = money_to_number(money_matches[-1]) if money_matches else ""

        for invoice in invoices:
            rows.append({
                "Supplier_name": "Prime",
                "Distname": "",
                "CustName": cust_name,
                "City": city,
                "State": state,
                "CustAccNbr": account_match.group(0) if account_match else "",
                "InvoiceNumber": invoice,
                "PO_Number": "",
                "UnitCost": "",
                "Qty": "",
                "Commissions": commission,
            })

    return rows


def extract_pdf(uploaded_file):
    rows = []

    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text:
                rows.extend(extract_rows_from_text(text))

    seen = set()
    unique_rows = []

    for row in rows:
        key = (
            row["InvoiceNumber"],
            row["CustAccNbr"],
            row["Commissions"],
        )
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    return unique_rows


def create_excel(rows):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

    ws.freeze_panes = "A2"
    wb.save(output)
    output.seek(0)
    return output


st.title("📊 Prime Conduit PDF Extractor")
st.write("Upload Prime Conduit Commission Report PDFs and download extracted invoice data as Excel.")

uploaded_files = st.file_uploader(
    "Upload PDF file(s)",
    type=["pdf"],
    accept_multiple_files=True
)

if uploaded_files:
    st.info(f"{len(uploaded_files)} file(s) uploaded.")

    if st.button("Extract Data", type="primary"):
        all_rows = []

        with st.spinner("Extracting invoice data..."):
            for uploaded_file in uploaded_files:
                try:
                    file_rows = extract_pdf(uploaded_file)
                    all_rows.extend(file_rows)
                    st.success(f"✅ {uploaded_file.name}: {len(file_rows)} rows extracted")
                except Exception as e:
                    st.error(f"❌ {uploaded_file.name}: {e}")

        if all_rows:
            st.subheader("Preview")
            st.dataframe(all_rows, use_container_width=True)

            excel_data = create_excel(all_rows)

            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_data,
                file_name="prime_conduit_extracted_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No invoice numbers found. Please check if the PDF text is selectable, not scanned image-only.")
else:
    st.warning("Please upload at least one PDF file.")
